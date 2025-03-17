import io
import os
import reportlab
import calendar

from django.db.models import Q
from django.http import HttpResponse
from reportlab.platypus import Table, TableStyle, Spacer, SimpleDocTemplate

from .generate_qr import generate_qrcode
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors, utils
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph
from reportlab.lib.utils import ImageReader
from reportlab.graphics.barcode import qr
from reportlab.graphics.shapes import Drawing
from reportlab.lib.units import cm, inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from schoolAssistance import settings

from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side

from PIL import Image as PILImage, ImageDraw
from datetime import datetime, timedelta

from ..school.models import School

styles = getSampleStyleSheet()
styles.add(
    ParagraphStyle(name='font-first-name', alignment=TA_CENTER, leading=8, fontName='NotoSerif-Bold', fontSize=11.8))
styles.add(
    ParagraphStyle(name='font-last-name', alignment=TA_CENTER, leading=8, fontName='NotoSerif-Regular', fontSize=6.5))
styles.add(ParagraphStyle(name='font-data', alignment=TA_LEFT, leading=8, fontName='NotoSerif-Regular', fontSize=7))
styles.add(ParagraphStyle(name='font-code', alignment=TA_CENTER, leading=8, fontName='NotoSerif-Regular', fontSize=5.5))

reportlab.rl_config.TTFSearchPath.append(str(settings.BASE_DIR) + '/static/fonts')
pdfmetrics.registerFont(TTFont('NotoSerif-Bold', 'NotoSerif-Bold.ttf'))
pdfmetrics.registerFont(TTFont('NotoSerif-Regular', 'NotoSerif-Regular.ttf'))
pdfmetrics.registerFont(TTFont('SourceSerifPro-Regular', 'SourceSerifPro-Regular.ttf'))


class FactoryFotocheck():
    get_name_grade = {
        '1': 'Primero',
        '2': 'Segundo',
        '3': 'Tercero',
        '4': 'Cuarto',
        '5': 'Quinto',
        '6': 'Sexto'
    }

    get_name_grade_inicial = {
        '1': '1 Año',
        '2': '2 Años',
        '3': '3 Años',
        '4': '4 Años',
        '5': '5 Años',
        '6': '6 Años'
    }

    def __init__(self):
        self.pagesize = A4
        self.margin_x = 20.00  # px
        self.margin_y = 20.00  # px
        self.usable_width = self.pagesize[0] - self.margin_y * 2  # arriba y abajo
        self.usable_height = self.pagesize[1] - self.margin_x * 2  # izquierda y derecha
        self.cell_width = self.usable_width / 3
        self.cell_height = self.usable_height / 3
        self.styles = {
            'Simple': ParagraphStyle(
                "Centrado",
                parent=getSampleStyleSheet()['Normal'],
                alignment=1,  # 1 es centrado
                fontSize=15,
                fontName='Helvetica-Bold',
                spaceAfter=100
            ),
            'SmallBlank': ParagraphStyle(
                "Centrado",
                parent=getSampleStyleSheet()['Normal'],
                alignment=1,  # 1 es centrado
                fontSize=10,
                fontName='Times-Roman',
                textColor=colors.white,
            )
        }

    def set_margin(self, new_margin_x=20, new_margin_y=20):
        self.margin_x = new_margin_x
        self.margin_y = new_margin_y
        self.usable_width = self.pagesize - self.margin_y * 2  # arriba y abajo
        self.usable_height = self.pagesize - self.margin_x * 2  # izquierda y derecha
        self.cell_width = self.usable_width / 3
        self.cell_height = self.usable_height / 3

    def create_qr(self, table):
        qr_code = qr.QrCodeWidget(table)
        bounds = qr_code.getBounds()
        width = bounds[2] - bounds[0]
        height = bounds[3] - bounds[1]
        drawing = Drawing(
            # 3.5 * cm, 3.5 * cm, transform=[3.5 * cm / width, 0, 0, 3.5 * cm / height, 0, 0])
            1 * cm, 1 * cm, transform=[4.5 * cm / width, 0, 0, 4.5 * cm / height, 0, 0])
        drawing.add(qr_code)
        return drawing

    def create_pdf_from_student(self):
        # Creacion del archivo
        buff = BytesIO()
        c = canvas.Canvas(buff, pagesize=self.pagesize)
        c.translate(self.margin_x, self.margin_y)
        # ------------------------------------------

        # Guardado----------------------------------
        c.showPage()
        c.save()
        buff.seek(0)
        return buff

    def encrypt_data(self, text, key):
        list_of_chars = ['á', 'Á', 'é', 'É', 'í', 'Í', 'ó', 'Ó', 'ú', 'Ú', 'ñ', 'Ñ']
        encrypted_text = ''
        for char in text:
            if char in list_of_chars:
                encrypted_text += char
            elif char.isalnum():
                if char.isalpha():
                    start = ord('a') if char.islower() else ord('A')
                    encrypted_text += chr((ord(char) - start + key) % 26 + start)
                else:
                    encrypted_text += chr((ord(char) - ord('0') + key) % 10 + ord('0'))
            else:
                encrypted_text += char

        return encrypted_text

    def get_file_path_from_url(self, url):
        path = url[len(settings.MEDIA_URL):]
        return os.path.join(settings.MEDIA_ROOT, path)

    def fit_name_on_image(self, first_name, max_length=18):
        first_name = first_name.upper()
        list_names = first_name.split()

        count = 0
        new_list_names = []

        for name in list_names:
            if count + len(name) + (1 if new_list_names else 0) <= max_length:
                new_list_names.append(name)
                count += len(name) + (1 if new_list_names else 0)
            else:
                break

        return ' '.join(new_list_names)

    def create_pdf_from_students(self, students_data, slug_cole):
        # Data students esta en formato diccionario, cada uno tiene
        # {"Nombre": "Juan Pedro", "Apellidos": "Pérez Gomez","Level": "Secundaria", "Grado":"5to", "Seccion": "A", "dni": 25425263, "Imagen": "fondo.png"},
        # Creacion del archivo
        buff = BytesIO()
        c = canvas.Canvas(buff, pagesize=self.pagesize)
        # c.translate(self.margin_x, self.margin_y)
        # ------------------------------------------
        n = len(students_data)
        static_dir = str(settings.BASE_DIR)

        school_obj = School.objects.get(slug=slug_cole)
        fotocheck_template_front_url = school_obj.fotocheck_template_front.url
        fotocheck_template_back_url = school_obj.fotocheck_template_back.url

        if n == 1:
            # Front
            # ***************************************************************************************************** #
            # ***************************************************************************************************** #

            id_cell = 0
            j = 0
            i = 0

            x = self.margin_x + j * self.cell_width
            y = self.margin_y + (2 * self.cell_height) - (i * self.cell_height)

            # Añadir imagen de fondo
            # background_path = os.path.join(static_dir, 'static', 'img', 'templates', f"{slug_cole}-front.png")
            background_path = self.get_file_path_from_url(fotocheck_template_front_url)
            profile_img_url = students_data[id_cell].get_profile_image()
            student_icon_path = static_dir + profile_img_url
            image = PILImage.open(student_icon_path)

            try:
                from .service_local import get_local_image
                image = get_local_image(students_data[id_cell].dni)
            except ModuleNotFoundError:
                pass

            #mask = PILImage.new("L", image.size, 0)
            #draw = ImageDraw.Draw(mask)
            #draw.ellipse((0, 0) + image.size, fill=255)

            #rounded_image = PILImage.new("RGBA", image.size)
            #rounded_image.paste(image, (0, 0), mask=mask)

            #rounded_image_reader = ImageReader(rounded_image)

            # logo_path = os.path.join(static_dir, 'img',"logo.png")
            c.drawImage(background_path, x, y, width=self.cell_width, height=self.cell_height,
                        preserveAspectRatio=True, anchor='c')
            # c.drawImage(student_icon_path, x + self.cell_width/4, y + self.cell_height/2, width=self.cell_width/3, height=self.cell_width/3, mask='auto')
            #c.drawImage(rounded_image_reader, x + self.cell_width / 3.1, y + self.cell_height / 1.735,
            #            width=self.cell_width / 2.82, height=self.cell_width / 2.82, mask='auto')

            # c.drawImage(logo_path, x+ self.cell_width - 45 , y+ self.cell_height - 50, width = 35, height= 40, mask='auto')
            # c.restoreState()
            first_name_data = self.fit_name_on_image(str(students_data[id_cell].first_name).upper())
            last_name_data = str(students_data[id_cell].last_name).upper()
            code = str(students_data[id_cell].dni)
            level_data = str(students_data[id_cell].level.name)
            grade_data = students_data[id_cell].grade.short_name
            section_data = students_data[id_cell].section.name

            if level_data.lower() == 'inicial':
                name_grade_data = self.get_name_grade_inicial[grade_data]
            else:
                name_grade_data = f'{self.get_name_grade[grade_data]} ({grade_data})'

            paragraph_first_name = Paragraph(first_name_data, styles['font-first-name'])
            paragraph_first_name.wrap(self.cell_width - 20,
                                      self.cell_height - 20)  # Ajustar el tamaño del área de texto
            paragraph_first_name.drawOn(c, x + 10, y + self.cell_height / 1.28)

            paragraph_last_name = Paragraph(last_name_data, styles['font-last-name'])
            paragraph_last_name.wrap(self.cell_width - 20, self.cell_height - 20)
            paragraph_last_name.drawOn(c, x + 10, y + self.cell_height / 1.38)

            #paragraph_code = Paragraph(code, styles['font-data'])
            #paragraph_code.wrap(self.cell_width - 20, self.cell_height - 20)
            #paragraph_code.drawOn(c, x + 93, y + self.cell_height / 2.06)

            paragraph_level = Paragraph(level_data, styles['font-data'])
            paragraph_level.wrap(self.cell_width - 20, self.cell_height - 20)
            paragraph_level.drawOn(c, x + 93, y + self.cell_height / 1.45)

            paragraph_grade = Paragraph(f'{name_grade_data}', styles['font-data'])
            paragraph_grade.wrap(self.cell_width - 20, self.cell_height - 20)
            paragraph_grade.drawOn(c, x + 93, y + self.cell_height / 1.55)

            paragraph_section = Paragraph(section_data, styles['font-data'])
            paragraph_section.wrap(self.cell_width - 20, self.cell_height - 20)
            paragraph_section.drawOn(c, x + 93, y + self.cell_height / 1.65)

            first_name_data = str(students_data[id_cell].first_name)
            last_name_data = str(students_data[id_cell].last_name)
            code = str(students_data[id_cell].dni)
            level_data = str(students_data[id_cell].level.name)
            grade_data = students_data[id_cell].grade.short_name
            section_data = students_data[id_cell].section.name

            data = f'{first_name_data}${last_name_data}${code}${grade_data}${section_data}'

            data_encrypt = self.encrypt_data(data, 3)

            qr_image_bytes = generate_qrcode(data=data_encrypt)
            qr_image = utils.ImageReader(BytesIO(qr_image_bytes))

            c.drawImage(qr_image, x + self.cell_width / 4.45, y + self.cell_height / 5,
                        width=self.cell_width / 2 + 10,
                        height=self.cell_width / 2 + 10, mask='auto')


            # Back
            # ***************************************************************************************************** #
            # ***************************************************************************************************** #
            i = 0
            j = 1

            x = self.margin_x + j * self.cell_width + 1
            y = self.margin_y + (2 * self.cell_height) - (i * self.cell_height)
            # Añadir imagen de fondo
            # background_path = os.path.join(static_dir, 'static', 'img', 'templates', f"{slug_cole}-back.png")
            background_path = self.get_file_path_from_url(fotocheck_template_back_url)
            # background_path = "back-bg.png"


            first_name_data = str(students_data[id_cell].first_name)
            last_name_data = str(students_data[id_cell].last_name)
            code = str(students_data[id_cell].dni)
            level_data = str(students_data[id_cell].level.name)
            grade_data = students_data[id_cell].grade.short_name
            section_data = students_data[id_cell].section.name

            data = f'{first_name_data}${last_name_data}${code}${grade_data}${section_data}'

            data_encrypt = self.encrypt_data(data, 3)

            qr_image_bytes = generate_qrcode(data=data_encrypt)
            qr_image = utils.ImageReader(BytesIO(qr_image_bytes))

            c.drawImage(background_path, x, y, width=self.cell_width, height=self.cell_height,
                        preserveAspectRatio=True, anchor='c')

            c.drawImage(qr_image, x + self.cell_width / 4.45, y + self.cell_height / 5,
                        width=self.cell_width / 2 + 10,
                        height=self.cell_width / 2 + 10, mask='auto')

            paragraph_code = Paragraph(code, styles['font-code'])
            paragraph_code.wrap(self.cell_width - 20, self.cell_height - 20)
            paragraph_code.drawOn(c, x + 10, y + self.cell_height / 5.5)

            c.showPage()
            c.save()
            buff.seek(0)

            return buff
        else:
            # Primera Página: Portada - Datos alumno
            for n_page in range(n // 9 + 1):  # para dividir por paginas, ya que entran 9 fotochecks por pagina
                for i, row in enumerate(range(3)):
                    for j, col in enumerate(range(3)):
                        id_cell = n_page * 9 + i * 3 + j
                        if (id_cell >= n):
                            break
                        x = self.margin_x + j * self.cell_width
                        y = self.margin_y + (2 * self.cell_height) - (i * self.cell_height)
                        # Añadir imagen de fondo
                        # background_path = os.path.join(static_dir, 'static', 'img', 'templates', f"{slug_cole}-front.png")
                        background_path = self.get_file_path_from_url(fotocheck_template_front_url)
                        profile_img_url = students_data[id_cell].get_profile_image()
                        student_icon_path = static_dir + profile_img_url
                        image = PILImage.open(student_icon_path)

                        try:
                            from .service_local import get_local_image
                            image = get_local_image(students_data[id_cell].dni)
                        except ModuleNotFoundError:
                            pass

                        mask = PILImage.new("L", image.size, 0)
                        draw = ImageDraw.Draw(mask)
                        draw.ellipse((0, 0) + image.size, fill=255)

                        rounded_image = PILImage.new("RGBA", image.size)
                        rounded_image.paste(image, (0, 0), mask=mask)

                        rounded_image_reader = ImageReader(rounded_image)

                        # logo_path = os.path.join(static_dir, 'img',"logo.png")
                        c.drawImage(background_path, x, y, width=self.cell_width, height=self.cell_height,
                                    preserveAspectRatio=True, anchor='c')
                        # c.drawImage(student_icon_path, x + self.cell_width/4, y + self.cell_height/2, width=self.cell_width/3, height=self.cell_width/3, mask='auto')
                        c.drawImage(rounded_image_reader, x + self.cell_width / 3.1, y + self.cell_height / 1.735,
                                    width=self.cell_width / 2.82, height=self.cell_width / 2.82, mask='auto')

                        # c.drawImage(logo_path, x+ self.cell_width - 45 , y+ self.cell_height - 50, width = 35, height= 40, mask='auto')
                        # c.restoreState()
                        first_name_data = self.fit_name_on_image(str(students_data[id_cell].first_name).upper())
                        last_name_data = str(students_data[id_cell].last_name).upper()
                        code = str(students_data[id_cell].dni)
                        level_data = str(students_data[id_cell].level.name)
                        grade_data = students_data[id_cell].grade.short_name
                        section_data = students_data[id_cell].section.name

                        if level_data.lower() == 'inicial':
                            name_grade_data = self.get_name_grade_inicial[grade_data]
                        else:
                            name_grade_data = f'{self.get_name_grade[grade_data]} ({grade_data})'

                        paragraph_first_name = Paragraph(first_name_data, styles['font-first-name'])
                        paragraph_first_name.wrap(self.cell_width - 20,
                                                  self.cell_height - 20)  # Ajustar el tamaño del área de texto
                        paragraph_first_name.drawOn(c, x + 10, y + self.cell_height / 1.95)

                        paragraph_last_name = Paragraph(last_name_data, styles['font-last-name'])
                        paragraph_last_name.wrap(self.cell_width - 20, self.cell_height - 20)
                        paragraph_last_name.drawOn(c, x + 10, y + self.cell_height / 2.18)

                        paragraph_code = Paragraph(code, styles['font-data'])
                        paragraph_code.wrap(self.cell_width - 20, self.cell_height - 20)
                        paragraph_code.drawOn(c, x + 93, y + self.cell_height / 2.575)

                        paragraph_level = Paragraph(level_data, styles['font-data'])
                        paragraph_level.wrap(self.cell_width - 20, self.cell_height - 20)
                        paragraph_level.drawOn(c, x + 93, y + self.cell_height / 2.85)

                        paragraph_grade = Paragraph(f'{name_grade_data}', styles['font-data'])
                        paragraph_grade.wrap(self.cell_width - 20, self.cell_height - 20)
                        paragraph_grade.drawOn(c, x + 93, y + self.cell_height / 3.18)

                        paragraph_section = Paragraph(section_data, styles['font-data'])
                        paragraph_section.wrap(self.cell_width - 20, self.cell_height - 20)
                        paragraph_section.drawOn(c, x + 93, y + self.cell_height / 3.60)

                c.showPage()
            # Segunda pagina: Reverso - Código QR
            for n_page in range(n // 9 + 1):  # para dividir por paginas, ya que entran 9 fotochecks por pagina
                for i, row in enumerate(range(3)):
                    for j, col in enumerate(range(3)):
                        id_cell = n_page * 9 + i * 3 + j
                        if (id_cell >= n):
                            break
                        x = self.margin_x + j * self.cell_width
                        y = self.margin_y + (2 * self.cell_height) - (i * self.cell_height)
                        # Añadir imagen de fondo
                        # background_path = os.path.join(static_dir, 'static', 'img', 'templates', f"{slug_cole}-back.png")
                        background_path = self.get_file_path_from_url(fotocheck_template_back_url)
                        # background_path = "back-bg.png"

                        first_name_data = str(students_data[id_cell].first_name)
                        last_name_data = str(students_data[id_cell].last_name)
                        code = str(students_data[id_cell].dni)
                        level_data = str(students_data[id_cell].level.name)
                        grade_data = students_data[id_cell].grade.short_name
                        section_data = students_data[id_cell].section.name

                        data = f'{first_name_data}${last_name_data}${code}${grade_data}${section_data}'

                        data_encrypt = self.encrypt_data(data, 3)

                        qr_image_bytes = generate_qrcode(data=data_encrypt)
                        qr_image = utils.ImageReader(BytesIO(qr_image_bytes))

                        c.drawImage(background_path, x, y, width=self.cell_width, height=self.cell_height,
                                    preserveAspectRatio=True, anchor='c')

                        c.drawImage(qr_image, x + self.cell_width / 4.45, y + self.cell_height / 2.56,
                                    width=self.cell_width / 2 + 10,
                                    height=self.cell_width / 2 + 10, mask='auto')

                        paragraph_code = Paragraph(code, styles['font-code'])
                        paragraph_code.wrap(self.cell_width - 20, self.cell_height - 20)
                        paragraph_code.drawOn(c, x + 10, y + self.cell_height / 3)
                c.showPage()
            # Guardado--------------------------------------
            c.showPage()
            c.save()
            buff.seek(0)

            return buff


    def create_pdf_from_student_card(self, students_data, slug_cole):
        # Data students esta en formato diccionario, cada uno tiene
        # {"Nombre": "Juan Pedro", "Apellidos": "Pérez Gomez","Level": "Secundaria", "Grado":"5to", "Seccion": "A", "dni": 25425263, "Imagen": "fondo.png"},
        # Creacion del archivo
        buff = BytesIO()
        c = canvas.Canvas(buff, pagesize=self.pagesize)
        # c.translate(self.margin_x, self.margin_y)
        # ------------------------------------------
        n = len(students_data)
        static_dir = str(settings.BASE_DIR)

        new_width = 5.2 * 28.35  # Carnet
        new_height = 8.34 * 28.35 # Carnet

        school_obj = School.objects.get(slug=slug_cole)
        fotocheck_template_front_url = school_obj.fotocheck_template_front.url
        fotocheck_template_back_url = school_obj.fotocheck_template_back.url

        if n == 1:
            # Front
            # ***************************************************************************************************** #
            # ***************************************************************************************************** #

            id_cell = 0
            j = 0
            i = 0

            x = self.margin_x + j * new_width
            y = self.margin_y + (2 * new_height) - (i * new_height) + 80

            # Añadir imagen de fondo
            # background_path = os.path.join(static_dir, 'static', 'img', 'templates', f"{slug_cole}-front.png")
            background_path = self.get_file_path_from_url(fotocheck_template_front_url)
            profile_img_url = students_data[id_cell].get_profile_image()
            student_icon_path = static_dir + profile_img_url
            image = PILImage.open(student_icon_path)

            try:
                from .service_local import get_local_image
                image = get_local_image(students_data[id_cell].dni)
            except ModuleNotFoundError:
                pass

            mask = PILImage.new("L", image.size, 0)
            draw = ImageDraw.Draw(mask)
            draw.ellipse((0, 0) + image.size, fill=255)

            rounded_image = PILImage.new("RGBA", image.size)
            rounded_image.paste(image, (0, 0), mask=mask)

            rounded_image_reader = ImageReader(rounded_image)

            # logo_path = os.path.join(static_dir, 'img',"logo.png")

            c.drawImage(background_path, x, y, width=new_width, height=new_height,
                        preserveAspectRatio=True, anchor='c')
            # c.drawImage(student_icon_path, x + self.cell_width/4, y + self.cell_height/2, width=self.cell_width/3, height=self.cell_width/3, mask='auto')
            c.drawImage(rounded_image_reader, x + new_width / 3.565, y + new_height / 1.7185,
                        width=new_width / 2.285, height=new_width / 2.285, mask='auto')

            # c.drawImage(logo_path, x+ self.cell_width - 45 , y+ self.cell_height - 50, width = 35, height= 40, mask='auto')
            # c.restoreState()
            first_name_data = self.fit_name_on_image(str(students_data[id_cell].first_name).upper(), max_length=17)
            last_name_data = str(students_data[id_cell].last_name).upper()
            code = str(students_data[id_cell].dni)
            level_data = str(students_data[id_cell].level.name)
            grade_data = students_data[id_cell].grade.short_name
            section_data = students_data[id_cell].section.name

            if level_data.lower() == 'inicial':
                name_grade_data = self.get_name_grade_inicial[grade_data]
            else:
                name_grade_data = f'{self.get_name_grade[grade_data]} ({grade_data})'

            paragraph_first_name = Paragraph(first_name_data, styles['font-first-name'])
            paragraph_first_name.wrap(new_width - 20,
                                      new_height - 20)  # Ajustar el tamaño del área de texto
            paragraph_first_name.drawOn(c, x + 10, y + new_height / 2)

            paragraph_last_name = Paragraph(last_name_data, styles['font-last-name'])
            paragraph_last_name.wrap(new_width - 20, new_height - 20)
            paragraph_last_name.drawOn(c, x + 10, y + new_height / 2.275)

            paragraph_code = Paragraph(code, styles['font-data'])
            paragraph_code.wrap(new_width - 20, new_height - 20)
            paragraph_code.drawOn(c, x + 75, y + new_height / 2.755)

            paragraph_level = Paragraph(level_data, styles['font-data'])
            paragraph_level.wrap(new_width - 20, new_height - 20)
            paragraph_level.drawOn(c, x + 75, y + new_height / 3.04)

            paragraph_grade = Paragraph(f'{name_grade_data}', styles['font-data'])
            paragraph_grade.wrap(new_width - 20, new_height - 20)
            paragraph_grade.drawOn(c, x + 75, y + new_height / 3.4)

            paragraph_section = Paragraph(section_data, styles['font-data'])
            paragraph_section.wrap(new_width - 20, new_height - 20)
            paragraph_section.drawOn(c, x + 75, y + new_height / 3.85)

            # Back
            # ***************************************************************************************************** #
            # ***************************************************************************************************** #
            i = 0
            j = 1

            x = self.margin_x + j * new_width + 0.5
            y = self.margin_y + (2 * new_height) - (i * new_height) + 80
            # Añadir imagen de fondo
            # background_path = os.path.join(static_dir, 'static', 'img', 'templates', f"{slug_cole}-back.png")
            background_path = self.get_file_path_from_url(fotocheck_template_back_url)
            # background_path = "back-bg.png"

            first_name_data = str(students_data[id_cell].first_name)
            last_name_data = str(students_data[id_cell].last_name)
            code = str(students_data[id_cell].dni)
            level_data = str(students_data[id_cell].level.name)
            grade_data = students_data[id_cell].grade.short_name
            section_data = students_data[id_cell].section.name

            data = f'{first_name_data}${last_name_data}${code}${grade_data}${section_data}'

            data_encrypt = self.encrypt_data(data, 3)

            qr_image_bytes = generate_qrcode(data=data_encrypt)
            qr_image = utils.ImageReader(BytesIO(qr_image_bytes))

            c.drawImage(background_path, x, y, width=new_width, height=new_height,
                        preserveAspectRatio=True, anchor='c')

            c.drawImage(qr_image, x + new_width / 5.25, y + new_height / 5,
                        width=new_width / 2 + 8,
                        height=new_width / 2 + 8, mask='auto')

            paragraph_code = Paragraph(code, styles['font-code'])
            paragraph_code.wrap(new_width - 20, new_height - 20)
            paragraph_code.drawOn(c, x + 4, y + new_height / 5.2)

            c.showPage()
            c.save()
            buff.seek(0)

            return buff
        else:
            # Primera Página: Portada - Datos alumno
            for n_page in range(n // 9 + 1):  # para dividir por paginas, ya que entran 9 fotochecks por pagina
                for i, row in enumerate(range(3)):
                    for j, col in enumerate(range(3)):
                        id_cell = n_page * 9 + i * 3 + j
                        if (id_cell >= n):
                            break
                        x = self.margin_x + j * new_width
                        y = self.margin_y + (2 * new_height) - (i * new_height)
                        # Añadir imagen de fondo
                        # background_path = os.path.join(static_dir, 'static', 'img', 'templates', f"{slug_cole}-front.png")
                        background_path = self.get_file_path_from_url(fotocheck_template_front_url)
                        profile_img_url = students_data[id_cell].get_profile_image()
                        student_icon_path = static_dir + profile_img_url
                        image = PILImage.open(student_icon_path)

                        try:
                            from .service_local import get_local_image
                            image = get_local_image(students_data[id_cell].dni)
                        except ModuleNotFoundError:
                            pass

                        mask = PILImage.new("L", image.size, 0)
                        draw = ImageDraw.Draw(mask)
                        draw.ellipse((0, 0) + image.size, fill=255)

                        rounded_image = PILImage.new("RGBA", image.size)
                        rounded_image.paste(image, (0, 0), mask=mask)

                        rounded_image_reader = ImageReader(rounded_image)

                        # logo_path = os.path.join(static_dir, 'img',"logo.png")
                        c.drawImage(background_path, x, y, width=new_width, height=new_height,
                                    preserveAspectRatio=True, anchor='c')
                        # c.drawImage(student_icon_path, x + self.cell_width/4, y + self.cell_height/2, width=self.cell_width/3, height=self.cell_width/3, mask='auto')
                        c.drawImage(rounded_image_reader, x + new_width / 3.1, y + new_height / 1.735,
                                    width=new_width / 2.82, height=new_width / 2.82, mask='auto')

                        # c.drawImage(logo_path, x+ self.cell_width - 45 , y+ self.cell_height - 50, width = 35, height= 40, mask='auto')
                        # c.restoreState()
                        first_name_data = self.fit_name_on_image(str(students_data[id_cell].first_name).upper())
                        last_name_data = str(students_data[id_cell].last_name).upper()
                        code = str(students_data[id_cell].dni)
                        level_data = str(students_data[id_cell].level.name)
                        grade_data = students_data[id_cell].grade.short_name
                        section_data = students_data[id_cell].section.name

                        if level_data.lower() == 'inicial':
                            name_grade_data = self.get_name_grade_inicial[grade_data]
                        else:
                            name_grade_data = f'{self.get_name_grade[grade_data]} ({grade_data})'

                        paragraph_first_name = Paragraph(first_name_data, styles['font-first-name'])
                        paragraph_first_name.wrap(new_width - 20,
                                                  new_height - 20)  # Ajustar el tamaño del área de texto
                        paragraph_first_name.drawOn(c, x + 10, y + new_height / 1.95)

                        paragraph_last_name = Paragraph(last_name_data, styles['font-last-name'])
                        paragraph_last_name.wrap(new_width - 20, new_height - 20)
                        paragraph_last_name.drawOn(c, x + 10, y + new_height / 2.18)

                        paragraph_code = Paragraph(code, styles['font-data'])
                        paragraph_code.wrap(new_width - 20, new_height - 20)
                        paragraph_code.drawOn(c, x + 93, y + new_height / 2.575)

                        paragraph_level = Paragraph(level_data, styles['font-data'])
                        paragraph_level.wrap(new_width - 20, new_height - 20)
                        paragraph_level.drawOn(c, x + 93, y + new_height / 2.85)

                        paragraph_grade = Paragraph(f'{name_grade_data}', styles['font-data'])
                        paragraph_grade.wrap(new_width - 20, new_height - 20)
                        paragraph_grade.drawOn(c, x + 93, y + new_height / 3.18)

                        paragraph_section = Paragraph(section_data, styles['font-data'])
                        paragraph_section.wrap(new_width - 20, new_height - 20)
                        paragraph_section.drawOn(c, x + 93, y + new_height / 3.60)

                c.showPage()
            # Segunda pagina: Reverso - Código QR
            for n_page in range(n // 9 + 1):  # para dividir por paginas, ya que entran 9 fotochecks por pagina
                for i, row in enumerate(range(3)):
                    for j, col in enumerate(range(3)):
                        id_cell = n_page * 9 + i * 3 + j
                        if (id_cell >= n):
                            break
                        x = self.margin_x + j * new_width
                        y = self.margin_y + (2 * new_height) - (i * new_height)
                        # Añadir imagen de fondo
                        # background_path = os.path.join(static_dir, 'static', 'img', 'templates', f"{slug_cole}-back.png")
                        background_path = self.get_file_path_from_url(fotocheck_template_back_url)
                        # background_path = "back-bg.png"

                        first_name_data = str(students_data[id_cell].first_name)
                        last_name_data = str(students_data[id_cell].last_name)
                        code = str(students_data[id_cell].dni)
                        level_data = str(students_data[id_cell].level.name)
                        grade_data = students_data[id_cell].grade.short_name
                        section_data = students_data[id_cell].section.name

                        data = f'{first_name_data}${last_name_data}${code}${grade_data}${section_data}'

                        data_encrypt = self.encrypt_data(data, 3)

                        qr_image_bytes = generate_qrcode(data=data_encrypt)
                        qr_image = utils.ImageReader(BytesIO(qr_image_bytes))

                        c.drawImage(background_path, x, y, width=new_width, height=new_height,
                                    preserveAspectRatio=True, anchor='c')

                        c.drawImage(qr_image, x + new_width / 4.45, y + new_height / 2.56,
                                    width=new_width / 2 + 10,
                                    height=new_width / 2 + 10, mask='auto')

                        paragraph_code = Paragraph(code, styles['font-code'])
                        paragraph_code.wrap(new_width - 20, new_height - 20)
                        paragraph_code.drawOn(c, x + 10, y + new_height / 3)
                c.showPage()
            # Guardado--------------------------------------
            c.showPage()
            c.save()
            buff.seek(0)

            return buff


    def create_qr_pdf_from_students(self, students_data, slug_cole):
        BASE = 21
        ALTURA = 29.7

        _wt = BASE * inch - 10 * 0.05 * inch

        ml = 0.0 * inch
        mr = 0.0 * inch
        ms = 0.039 * inch
        mi = 0.039 * inch

        _dictionary = []

        students = students_data

        style_table_1 = [
            ('BOX', (0, 0), (-1, -1), 2, colors.black),
            ('BOX', (0, 1), (-1, -1), 2, colors.black),
            ('SPAN', (0, 0), (-1, 0)),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (0, 0), -15),
            ('LEFTPADDING', (0, 0), (0, 0), 60),
            ('LEFTPADDING', (0, 1), (0, -1), 20),

        ]

        lenght_students = len(students)

        ana_c1 = ''
        ana_c2 = ''
        ana_c3 = ''

        if True:
            for i in range(0, lenght_students, 3):
                if i < lenght_students:
                    student = students[i]

                    first_name_data = str(student.first_name)
                    last_name_data = str(student.last_name)
                    code = str(student.dni)
                    level_data = str(student.level.name)
                    grade_data = student.grade.short_name
                    section_data = student.section.name

                    data = f'{first_name_data}${last_name_data}${code}${grade_data}${section_data}'

                    data_encrypt = self.encrypt_data(data, 3)

                    qr_image_bytes = generate_qrcode(data=data_encrypt)
                    qr_image = utils.ImageReader(BytesIO(qr_image_bytes))

                    image_data = Image(qr_image)

                    image_data.drawHeight = inch * 1
                    image_data.drawWidth = inch * 1

                    p1_1 = Paragraph(f'Nombre:', styles["Right"])
                    p1_2 = Paragraph(f'{first_name_data}', styles["Left"])
                    p1_3 = Paragraph(f'Apellido:', styles["Right"])
                    p1_4 = Paragraph(f'{last_name_data}', styles["Left"])
                    p1_5 = Paragraph(f'DNI:', styles["Right"])
                    p1_6 = Paragraph(f'{code}', styles["Left"])
                    p1_7 = Paragraph(f'Grado:', styles["Right"])
                    p1_8 = Paragraph(f'{grade_data}', styles["Left"])
                    p1_9 = Paragraph(f'Sección:', styles["Right"])
                    p1_10 = Paragraph(f'{section_data}', styles["Left"])

                    colwiths_table_1 = [_wt * 40 / 100 * 0.32, _wt * 5 / 100 * 0.32, _wt * 55 / 100 * 0.32]
                    rowwiths_table_1 = [inch * 4, inch * 0.5, inch * 0.5, inch * 0.5, inch * 0.5, inch * 0.5]
                    ana_c1 = Table(
                        [(image_data, '', '')] +
                        [(p1_1, '', p1_2)] +
                        [(p1_3, '', p1_4)] +
                        [(p1_5, '', p1_6)] +
                        [(p1_7, '', p1_8)] +
                        [(p1_9, '', p1_10)],
                        colWidths=colwiths_table_1, rowHeights=rowwiths_table_1)
                    ana_c1.setStyle(TableStyle(style_table_1))

                # -------------------------------------------------------------------------------------------------- #
                # -------------------------------------------------------------------------------------------------- #
                if i + 1 < lenght_students:
                    student = students[i + 1]
                    first_name_data = str(student.first_name)
                    last_name_data = str(student.last_name)
                    code = str(student.dni)
                    level_data = str(student.level.name)
                    grade_data = student.grade.short_name
                    section_data = student.section.name

                    data = f'{first_name_data}${last_name_data}${code}${grade_data}${section_data}'

                    data_encrypt = self.encrypt_data(data, 3)

                    qr_image_bytes = generate_qrcode(data=data_encrypt)
                    qr_image = utils.ImageReader(BytesIO(qr_image_bytes))

                    image_data = Image(qr_image)

                    image_data.drawHeight = inch * 1
                    image_data.drawWidth = inch * 1

                    p2_1 = Paragraph(f'Nombre:', styles["Right"])
                    p2_2 = Paragraph(f'{first_name_data}', styles["Left"])
                    p2_3 = Paragraph(f'Apellido:', styles["Right"])
                    p2_4 = Paragraph(f'{last_name_data}', styles["Left"])
                    p2_5 = Paragraph(f'DNI:', styles["Right"])
                    p2_6 = Paragraph(f'{code}', styles["Left"])
                    p2_7 = Paragraph(f'Grado:', styles["Right"])
                    p2_8 = Paragraph(f'{grade_data}', styles["Left"])
                    p2_9 = Paragraph(f'Sección:', styles["Right"])
                    p2_10 = Paragraph(f'{section_data}', styles["Left"])

                    colwiths_table_2 = [_wt * 40 / 100 * 0.32, _wt * 5 / 100 * 0.32, _wt * 55 / 100 * 0.32]
                    rowwiths_table_2 = [inch * 4, inch * 0.5, inch * 0.5, inch * 0.5, inch * 0.5, inch * 0.5]
                    ana_c2 = Table(
                        [(image_data, '', '')] +
                        [(p2_1, '', p2_2)] +
                        [(p2_3, '', p2_4)] +
                        [(p2_5, '', p2_6)] +
                        [(p2_7, '', p2_8)] +
                        [(p2_9, '', p2_10)],
                        colWidths=colwiths_table_2, rowHeights=rowwiths_table_2)
                    ana_c2.setStyle(TableStyle(style_table_1))

                # -------------------------------------------------------------------------------------------------- #
                # -------------------------------------------------------------------------------------------------- #
                if i + 2 < lenght_students:
                    student = students[i + 2]
                    first_name_data = str(student.first_name)
                    last_name_data = str(student.last_name)
                    code = str(student.dni)
                    level_data = str(student.level.name)
                    grade_data = student.grade.short_name
                    section_data = student.section.name

                    data = f'{first_name_data}${last_name_data}${code}${grade_data}${section_data}'

                    data_encrypt = self.encrypt_data(data, 3)

                    qr_image_bytes = generate_qrcode(data=data_encrypt)
                    qr_image = utils.ImageReader(BytesIO(qr_image_bytes))

                    image_data = Image(qr_image)

                    image_data.drawHeight = inch * 1
                    image_data.drawWidth = inch * 1

                    p3_1 = Paragraph(f'Nombre:', styles["Right"])
                    p3_2 = Paragraph(f'{first_name_data}', styles["Left"])
                    p3_3 = Paragraph(f'Apellido:', styles["Right"])
                    p3_4 = Paragraph(f'{last_name_data}', styles["Left"])
                    p3_5 = Paragraph(f'DNI:', styles["Right"])
                    p3_6 = Paragraph(f'{code}', styles["Left"])
                    p3_7 = Paragraph(f'Grado:', styles["Right"])
                    p3_8 = Paragraph(f'{grade_data}', styles["Left"])
                    p3_9 = Paragraph(f'Sección:', styles["Right"])
                    p3_10 = Paragraph(f'{section_data}', styles["Left"])

                    colwiths_table_3 = [_wt * 40 / 100 * 0.32, _wt * 5 / 100 * 0.32, _wt * 55 / 100 * 0.32]
                    rowwiths_table_3 = [inch * 4, inch * 0.5, inch * 0.5, inch * 0.5, inch * 0.5, inch * 0.5]
                    ana_c3 = Table(
                        [(image_data, '', '')] +
                        [(p3_1, '', p3_2)] +
                        [(p3_3, '', p3_4)] +
                        [(p3_5, '', p3_6)] +
                        [(p3_7, '', p3_8)] +
                        [(p3_9, '', p3_10)],
                        colWidths=colwiths_table_3, rowHeights=rowwiths_table_3)
                    ana_c3.setStyle(TableStyle(style_table_1))

                # -------------------------------------------------------------------------------------------------- #
                # -------------------------------------------------------------------------------------------------- #

                style_table_4 = [
                    ('TOPPADDING', (0, 0), (-1, -1), 0),
                    ('LEFTPADDING', (0, 0), (-1, -1), 0),
                    ('RIGHPADDING', (0, 0), (-1, -1), 0),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                ]

                colwiths_table_4 = [_wt * 32 / 100, _wt * 2 / 100, _wt * 32 / 100, _wt * 2 / 100, _wt * 32 / 100]
                # rowwiths_table_4 = [inch * 1]
                ana_c4 = Table(
                    [(ana_c1, '', ana_c2, '', ana_c3)],
                    # colWidths=colwiths_table_4, rowHeights=rowwiths_table_4)
                    colWidths=colwiths_table_4)
                ana_c4.setStyle(TableStyle(style_table_4))

                _dictionary.append(Spacer(width=8, height=16))
                _dictionary.append(ana_c4)

        buff = io.BytesIO()

        pz_matricial = (2.57 * inch, 11.6 * inch)
        # pz_termical = (3.14961 * inch, 11.6 * inch)
        pz_termical = (BASE * inch, ALTURA * inch)

        doc = SimpleDocTemplate(buff,
                                pagesize=pz_termical,
                                rightMargin=mr,
                                leftMargin=ml,
                                topMargin=ms,
                                bottomMargin=mi,
                                title='TICKET'
                                )
        doc.build(_dictionary)
        # doc.build(elements)
        # doc.build(Story)
        #
        # response = HttpResponse(content_type='application/pdf')
        # response['Content-Disposition'] = 'attachment; filename="{}-{}.pdf"'.format(order_obj.nombres, order_obj.pagos.id)
        #

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="[{}].pdf"'.format('qr' + ' - ' + 'file')

        tomorrow = datetime.now() + timedelta(days=1)
        tomorrow = tomorrow.replace(hour=0, minute=0, second=0)
        expires = datetime.strftime(tomorrow, "%a, %d-%b-%Y %H:%M:%S GMT")

        response.set_cookie('bp', value=1, expires=expires)

        response.write(buff.getvalue())

        buff.close()
        return response


class FactoryReports():
    def __init__(self):
        self.name_report = "Reporte de Aistencia General"

    def get_month(self, month_number):
        months = {
            1: "Enero",
            2: "Febrero",
            3: "Marzo",
            4: "Abril",
            5: "Mayo",
            6: "Junio",
            7: "Julio",
            8: "Agosto",
            9: "Septiembre",
            10: "Octubre",
            11: "Noviembre",
            12: "Diciembre"
        }

        return months[month_number]

    def get_icon_report(self, status):
        icon = {
            'Presente': '✅',
            'Falta': '❌',
            'Tardanza': 'T',
            'Desconocido': '',
            'Tardanza Justificada Pedida': 'J',
            'Tardanza Justificada Registrada': 'J',
            'Falta Justificada': 'J',
            'Permiso': 'J'
        }
        return icon[status]

    def calculate_percentage(self, count, total_count):
        if total_count == 0:
            return 0
        percentage = (count / total_count) * 100
        if percentage.is_integer():
            return int(percentage)
        else:
            return round(percentage, 1)

    def create_general_report(self, students_data, user_request, context):
        register_assistance_path = os.path.join(settings.BASE_DIR, 'static', 'resources', 'registro-asistencia.xlsx')
        book = openpyxl.load_workbook(register_assistance_path)
        sheet = book.active

        logo_media_path = context["school"].get_logo()
        logo_path = str(settings.BASE_DIR) + logo_media_path

        # Resize image
        try:
            img = PILImage.open(logo_path)
            original_width, original_height = img.size
            new_height = 140
            aspect_ratio = original_width / original_height
            new_width = int(new_height * aspect_ratio)

            new_size = (new_width, new_height)
            img = img.resize(new_size)

            resized_logo_dir = os.path.join(settings.BASE_DIR, 'media', 'resized_logo')
            resized_logo_path = os.path.join(resized_logo_dir, f'{context["school"].slug}-resized.png')

            if not os.path.exists(resized_logo_dir):
                os.makedirs(resized_logo_dir)

            img.save(resized_logo_path)

            img = Image(resized_logo_path)
            sheet.add_image(img, 'C3')

        except Exception as e:
            print(f"Error procesando la imagen del logo: {e}")

        # School
        sheet['F4'] = f'{context["school"].name}'

        # User
        sheet['F5'] = f'{user_request}'

        # Level
        sheet['F6'] = f'{context["level"]}'

        # Grade
        sheet['F7'] = f'{context["grade"]}'

        # Section
        sheet['F8'] = f'{context["section"]}'

        # Year
        sheet['W7'] = f'{datetime.now().year}'
        # sheet['W7'].font = Font(color='FFFFFF', bold=True)
        # sheet['W7'].alignment = Alignment(horizontal='center', vertical='center')

        # Month
        month_name = self.get_month(context["month"])
        sheet['W8'] = f'{month_name}'
        # sheet['W8'].alignment = Alignment(horizontal='center', vertical='center')

        # Department
        sheet['F9'] = f'{context["school"].department}'
        # sheet['F9'].font = Font(bold=True)

        # Province
        sheet['N9'] = f'{context["school"].province}'
        # sheet['N9'].font = Font(bold=True)

        # District
        sheet['V9'] = f'{context["school"].district}'
        # sheet['V9'].font = Font(bold=True)

        # month numbers
        week_days = {
            0: 'L',
            1: 'M',
            2: 'M',
            3: 'J',
            4: 'V',
            5: 'S',
            6: 'D'
        }

        current_year = datetime.now().year
        current_month = context['month']
        current_calendar = calendar.monthcalendar(current_year, current_month)

        column_name_day = 6  # Columna F
        row_name_day = 11  # Fila 11

        column_day = 6  # Columna F
        row_day = 12  # Fila 12

        for week_i in current_calendar:
            for day in week_i:
                if day != 0:
                    current_date = datetime(current_year, current_month, day)

                    week_day = current_date.weekday()
                    name_day = week_days[week_day]

                    sheet.cell(row=row_name_day, column=column_name_day, value=name_day)
                    sheet.cell(row=row_day, column=column_day, value=day)

                    column_name_day += 1
                    column_day += 1

        # Students
        align_center = Alignment(horizontal='center')

        # Start students cell B13 - C13 - F13
        column_student_code = 2
        row_student_code = 13

        column_student_name = 3
        row_student_name = 13

        column_student_assistance = 6
        row_student_assistance = 13

        fill_tardanza = PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')
        fill_justificacion = PatternFill(start_color='ff9900', end_color='ff9900', fill_type='solid')
        fill_desconocido = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        fill_presente = PatternFill(start_color='00d26a', end_color='00d26a', fill_type='solid')
        fill_falta = PatternFill(start_color='f92f60', end_color='f92f60', fill_type='solid')
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

        # Counts
        presente_count_total = 0
        falta_count_total = 0
        justificacion_count_total = 0  # justificacion_count_total = "Tardanza Justificada Pedida" + 'Tardanza Justificada Registrada' +  'Falta Justificada'
        tardanza_count_total = 0

        for row_num, student_data in enumerate(students_data):

            # Student Code
            code_cell = sheet.cell(row=row_student_code, column=column_student_code, value=f'{student_data.dni}')
            code_cell.alignment = align_center
            code_cell.border = border
            # row_student_code += 1

            # Student Full Name
            name_cell = sheet.cell(row=row_student_name, column=column_student_name,
                                   value=f'{student_data.last_name}, {student_data.first_name}')
            name_cell.border = border

            sheet.merge_cells(start_row=row_student_name, start_column=column_student_name, end_row=row_student_name,
                              end_column=column_student_name + 1)

            row_student_name += 1

            # Student Assistance
            assistances_in_month_and_year = student_data.general_assistances.filter(
                Q(general_assistance__date__month=current_month) &
                Q(general_assistance__date__year=current_year)
            )

            # Counts
            presente_count_student = 0
            falta_count_student = 0
            justificacion_count_student = 0  # justificacion_count_total = "Tardanza Justificada Pedida" + 'Tardanza Justificada Registrada' +  'Falta Justificada'
            tardanza_count_student = 0

            for assistance in assistances_in_month_and_year:
                icon_report = self.get_icon_report(assistance.state)
                cell_day = sheet.cell(row=row_student_assistance, column=column_student_assistance,
                                      value=f'{icon_report}')
                if icon_report == 'T':
                    cell_day.fill = fill_tardanza
                    cell_day.font = Font(bold=True)
                    cell_day.alignment = Alignment(horizontal='center', vertical='center')

                    tardanza_count_student += 1

                elif icon_report == 'J':
                    cell_day.fill = fill_justificacion
                    cell_day.font = Font(bold=True)
                    cell_day.alignment = Alignment(horizontal='center', vertical='center')

                    justificacion_count_student += 1

                elif assistance.state == 'Desconocido':
                    cell_day.fill = fill_desconocido
                elif assistance.state == 'Presente':
                    cell_day.fill = fill_presente

                    presente_count_student += 1

                elif assistance.state == 'Falta':
                    cell_day.fill = fill_falta

                    falta_count_student += 1

                cell_day.border = border

                column_student_assistance += 1

            column_assistance = 38

            count_presente_cell = sheet.cell(row=row_student_code, column=column_assistance,
                                             value=f'{presente_count_student}')
            count_falta_cell = sheet.cell(row=row_student_code, column=column_assistance + 1,
                                          value=f'{falta_count_student}')
            count_tardanza_cell = sheet.cell(row=row_student_code, column=column_assistance + 2,
                                             value=f'{tardanza_count_student}')
            count_justificacion_cell = sheet.cell(row=row_student_code, column=column_assistance + 3,
                                                  value=f'{justificacion_count_student}')

            count_presente_cell.border = border
            count_presente_cell.alignment = Alignment(horizontal='center', vertical='center')

            count_falta_cell.border = border
            count_falta_cell.alignment = Alignment(horizontal='center', vertical='center')

            count_tardanza_cell.border = border
            count_tardanza_cell.alignment = Alignment(horizontal='center', vertical='center')

            count_justificacion_cell.border = border
            count_justificacion_cell.alignment = Alignment(horizontal='center', vertical='center')

            row_student_code += 1

            presente_count_total += presente_count_student
            falta_count_total += falta_count_student
            justificacion_count_total += justificacion_count_student
            tardanza_count_total += tardanza_count_student

            column_student_assistance = 6
            row_student_assistance += 1

        total_count = presente_count_total + falta_count_total + justificacion_count_total + tardanza_count_total
        presente_percentage = self.calculate_percentage(presente_count_total, total_count)
        falta_percentage = self.calculate_percentage(falta_count_total, total_count)
        tardanza_percentage = self.calculate_percentage(tardanza_count_total, total_count)
        justificacion_percentage = self.calculate_percentage(justificacion_count_total, total_count)

        column_percentage = 35
        row_percentage = 5

        count_presente_cell = sheet.cell(row=row_percentage, column=column_percentage, value=f'{presente_percentage} %')
        count_presente_cell = sheet.cell(row=row_percentage + 1, column=column_percentage,
                                         value=f'{falta_percentage} %')
        count_presente_cell = sheet.cell(row=row_percentage + 2, column=column_percentage,
                                         value=f'{tardanza_percentage} %')
        count_presente_cell = sheet.cell(row=row_percentage + 3, column=column_percentage,
                                         value=f'{justificacion_percentage} %')

        # book.save('test-registro-asistencia.xlsx')
        return book

    def create_course_report(self, students_data, user_request, context):
        register_assistance_path = os.path.join(settings.BASE_DIR, 'static', 'resources',
                                                'registro-asistencia-curso.xlsx')
        book = openpyxl.load_workbook(register_assistance_path)
        sheet = book.active

        logo_media_path = context["school"].get_logo()
        logo_path = str(settings.BASE_DIR) + logo_media_path

        # Resize image
        try:
            img = PILImage.open(logo_path)
            original_width, original_height = img.size
            new_height = 140
            aspect_ratio = original_width / original_height
            new_width = int(new_height * aspect_ratio)

            new_size = (new_width, new_height)
            img = img.resize(new_size)

            resized_logo_dir = os.path.join(settings.BASE_DIR, 'media', 'resized_logo')
            resized_logo_path = os.path.join(resized_logo_dir, f'{context["school"].slug}-resized.png')

            if not os.path.exists(resized_logo_dir):
                os.makedirs(resized_logo_dir)

            img.save(resized_logo_path)

            img = Image(resized_logo_path)
            sheet.add_image(img, 'C3')

        except Exception as e:
            print(f"Error procesando la imagen del logo: {e}")

        # School
        sheet['F4'] = f'{context["school"].name}'

        # User
        sheet['F5'] = f'{user_request}'

        # Level
        sheet['F6'] = f'{context["level"]}'

        # Grade
        sheet['F7'] = f'{context["grade"]}'

        # Section
        sheet['F8'] = f'{context["section"]}'

        # Course
        sheet['N8'] = f'{context["course"].name}'

        # Year
        sheet['W7'] = f'{datetime.now().year}'
        # sheet['W7'].font = Font(color='FFFFFF', bold=True)
        # sheet['W7'].alignment = Alignment(horizontal='center', vertical='center')

        # Month
        month_name = self.get_month(context["month"])
        sheet['W8'] = f'{month_name}'
        # sheet['W8'].alignment = Alignment(horizontal='center', vertical='center')

        # Department
        sheet['F9'] = f'{context["school"].department}'
        # sheet['F9'].font = Font(bold=True)

        # Province
        sheet['N9'] = f'{context["school"].province}'
        # sheet['N9'].font = Font(bold=True)

        # District
        sheet['V9'] = f'{context["school"].district}'
        # sheet['V9'].font = Font(bold=True)

        # month numbers
        week_days = {
            0: 'L',
            1: 'M',
            2: 'M',
            3: 'J',
            4: 'V',
            5: 'S',
            6: 'D'
        }

        current_year = datetime.now().year
        current_month = context['month']
        current_calendar = calendar.monthcalendar(current_year, current_month)

        column_name_day = 6  # Columna F
        row_name_day = 11  # Fila 11

        column_day = 6  # Columna F
        row_day = 12  # Fila 12

        for week_i in current_calendar:
            for day in week_i:
                if day != 0:
                    current_date = datetime(current_year, current_month, day)

                    week_day = current_date.weekday()
                    name_day = week_days[week_day]

                    sheet.cell(row=row_name_day, column=column_name_day, value=name_day)
                    sheet.cell(row=row_day, column=column_day, value=day)

                    column_name_day += 1
                    column_day += 1

        # Students
        align_center = Alignment(horizontal='center')

        # Start students cell B13 - C13 - F13
        column_student_code = 2
        row_student_code = 13

        column_student_name = 3
        row_student_name = 13

        column_student_assistance = 6
        row_student_assistance = 13

        fill_tardanza = PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')
        fill_justificacion = PatternFill(start_color='ff9900', end_color='ff9900', fill_type='solid')
        fill_desconocido = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        fill_presente = PatternFill(start_color='00d26a', end_color='00d26a', fill_type='solid')
        fill_falta = PatternFill(start_color='f92f60', end_color='f92f60', fill_type='solid')
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

        # Counts
        presente_count_total = 0
        falta_count_total = 0
        justificacion_count_total = 0  # justificacion_count_total = "Tardanza Justificada Pedida" + 'Tardanza Justificada Registrada' +  'Falta Justificada'
        tardanza_count_total = 0

        for row_num, student_data in enumerate(students_data):

            # Student Code
            code_cell = sheet.cell(row=row_student_code, column=column_student_code, value=f'{student_data.dni}')
            code_cell.alignment = align_center
            code_cell.border = border
            # row_student_code += 1

            # Student Full Name
            name_cell = sheet.cell(row=row_student_name, column=column_student_name,
                                   value=f'{student_data.last_name}, {student_data.first_name}')
            name_cell.border = border

            sheet.merge_cells(start_row=row_student_name, start_column=column_student_name, end_row=row_student_name,
                              end_column=column_student_name + 1)

            row_student_name += 1

            # Student Assistance
            assistances_in_month_and_year = student_data.assistances.filter(
                Q(assistance__date__month=current_month) &
                Q(assistance__date__year=current_year) &
                Q(assistance__course=context['course'])
            )

            # Counts
            presente_count_student = 0
            falta_count_student = 0
            justificacion_count_student = 0
            tardanza_count_student = 0
            column_day = 5
            for assistance in assistances_in_month_and_year:
                column_day += assistance.assistance.date.day

                icon_report = self.get_icon_report(assistance.state)
                cell_day = sheet.cell(row=row_student_assistance, column=column_day,
                                      value=f'{icon_report}')
                if icon_report == 'T':
                    cell_day.fill = fill_tardanza
                    cell_day.font = Font(bold=True)
                    cell_day.alignment = Alignment(horizontal='center', vertical='center')

                    tardanza_count_student += 1

                elif icon_report == 'J':
                    cell_day.fill = fill_justificacion
                    cell_day.font = Font(bold=True)
                    cell_day.alignment = Alignment(horizontal='center', vertical='center')

                    justificacion_count_student += 1

                elif assistance.state == 'Desconocido':
                    # cell_day.fill = fill_desconocido
                    # cell_day.fill = fill_falta
                    #
                    # falta_count_student += 1
                    pass

                elif assistance.state == 'Presente':
                    cell_day.fill = fill_presente

                    presente_count_student += 1

                elif assistance.state == 'Falta':
                    cell_day.fill = fill_falta

                    falta_count_student += 1

                cell_day.border = border

                column_day = 5
                column_student_assistance += 1

            column_assistance = 38

            count_presente_cell = sheet.cell(row=row_student_code, column=column_assistance,
                                             value=f'{presente_count_student}')
            count_falta_cell = sheet.cell(row=row_student_code, column=column_assistance + 1,
                                          value=f'{falta_count_student}')
            count_tardanza_cell = sheet.cell(row=row_student_code, column=column_assistance + 2,
                                             value=f'{tardanza_count_student}')
            count_justificacion_cell = sheet.cell(row=row_student_code, column=column_assistance + 3,
                                                  value=f'{justificacion_count_student}')

            count_presente_cell.border = border
            count_presente_cell.alignment = Alignment(horizontal='center', vertical='center')

            count_falta_cell.border = border
            count_falta_cell.alignment = Alignment(horizontal='center', vertical='center')

            count_tardanza_cell.border = border
            count_tardanza_cell.alignment = Alignment(horizontal='center', vertical='center')

            count_justificacion_cell.border = border
            count_justificacion_cell.alignment = Alignment(horizontal='center', vertical='center')

            row_student_code += 1

            presente_count_total += presente_count_student
            falta_count_total += falta_count_student
            justificacion_count_total += justificacion_count_student
            tardanza_count_total += tardanza_count_student

            column_student_assistance = 6
            row_student_assistance += 1

        total_count = presente_count_total + falta_count_total + justificacion_count_total + tardanza_count_total
        presente_percentage = self.calculate_percentage(presente_count_total, total_count)
        falta_percentage = self.calculate_percentage(falta_count_total, total_count)
        tardanza_percentage = self.calculate_percentage(tardanza_count_total, total_count)
        justificacion_percentage = self.calculate_percentage(justificacion_count_total, total_count)

        column_percentage = 35
        row_percentage = 5

        count_presente_cell = sheet.cell(row=row_percentage, column=column_percentage, value=f'{presente_percentage} %')
        count_presente_cell = sheet.cell(row=row_percentage + 1, column=column_percentage,
                                         value=f'{falta_percentage} %')
        count_presente_cell = sheet.cell(row=row_percentage + 2, column=column_percentage,
                                         value=f'{tardanza_percentage} %')
        count_presente_cell = sheet.cell(row=row_percentage + 3, column=column_percentage,
                                         value=f'{justificacion_percentage} %')

        # book.save('test-registro-asistencia.xlsx')
        return book

    # def create_course_report(self, students_data, course, assistances_course_month, month):
    #     # Crear un nuevo libro de trabajo
    #     libro_trabajo = openpyxl.Workbook()
    #
    #     # Seleccionar la hoja activa
    #     hoja_activa = libro_trabajo.active
    #     hoja_activa.column_dimensions.default_width = 20
    #     hoja_activa.row_dimensions.default_height = 10
    #     # Escribir encabezados
    #     hoja_activa['A1'] = 'Nombre'
    #     hoja_activa.column_dimensions['A'].width = 25
    #     hoja_activa.row_dimensions[1].height = 12
    #     hoja_activa['B1'] = 'Clase'
    #
    #     init_x = 2
    #     init_y = 3
    #
    #     for col_num, course in enumerate(assistances_course_month):  # Columnas B a K
    #         hoja_activa.cell(row=1, column=init_x + col_num,
    #                          value=f'{course.name}\n{course.schedule_init} - {course.schedule_end}').alignment = Alignment(
    #             wrap_text=True)
    #     # Escribir datos en celdas
    #     for row_num, student_data in enumerate(students_data, start=1):
    #         hoja_activa.cell(row=row_num + init_x, column=1,
    #                          value=f'{student_data.first_name} {student_data.last_name}')
    #         for col_num, assistance in enumerate(student_data.assistances.all()):
    #             if assistance.assistance.date.month == month and assistance.assistance.course == course:
    #                 if (assistance.state == "Presente" or assistance.state == "Tardanza"):
    #                     hoja_activa.cell(row=row_num + init_x, column=col_num + init_y + 1, value=f'{assistance.time}')
    #                 elif (assistance.state == "Falta"):
    #                     hoja_activa.cell(row=row_num + init_x, column=col_num + init_y + 1, value="Falta")
    #                 else:
    #                     hoja_activa.cell(row=row_num + init_x, column=col_num + init_y + 1, value="None")
    #
    #     # Guardar el libro de trabajo en un archivo
    #     return libro_trabajo
